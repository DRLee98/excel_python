from openpyxl import load_workbook, Workbook

# 한글 인코딩 문제 해결
def format_text(value):
    if(type(value) != int):
        return value.encode('ISO-8859-1').decode('cp949')
    else:
        return value

# 특정 컬럼의 인덱스를 반환하는 함수
def get_target_col(ws, target_string):
    target_col = -1;
    for i in range(ws.max_column):
        cell = ws.cell(row=1, column=i+1)
        if(format_text(cell.value) == target_string):
            target_col = cell.col_idx
            break

    return target_col

# 엑셀 파일의 특정 컬럼에서 특정 문자열을 포함하는 행과 그 행의 회사명을 반환하는 함수
def find_row_and_company_name(file_name, col_target_name, target_string):
    wb = load_workbook(file_name, data_only=True)
    ws = wb[wb.sheetnames[0]]

    row_nums = []
    header = []
    result = []

    target_col = get_target_col(ws, col_target_name)

    # 타겟 컬럼이 없을 경우
    if(target_col == -1):
        print("No target column")
        return result

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=target_col, max_col=target_col):
        for cell in row:
            # 타겟 문자열이 포함된 행을 탐색
            if(target_string in format_text(cell.value).replace(" " , "")):
                row_nums.append(cell.row)

    company_name_col = get_target_col(ws, '회사명')
    for row_num in row_nums:
        row = ws[row_num]
        company_name = format_text(row[company_name_col].value)
        result.append({'row': row, 'name': company_name})
    
    header = ws[1]

    wb.close()

    return {"list": result, "header": header}

# 기준 리스트를 기준으로 옵션 리스트와 회사명을 비교하여 일치하는 목록을 반환하는 함수
def compare_company_names(criteria_list, option_list):
    copy_option_list = option_list.copy()
    result = []
    for item in criteria_list:
        list = [item]
        for option in copy_option_list:
            if(item['name'] == option['name']):
                list.append(option)
        
        if(len(list) > 1):
            result = result + list
    
    # 딕셔너리를 frozenset으로 변환하여 중복 제거
    unique_result = {frozenset(item.items()) for item in result}
    # 중복이 제거된 딕셔너리 리스트 생성
    unique_result_list = [dict(items) for items in unique_result]

    sorted_list = sorted(unique_result_list, key=lambda x: x['name'])

    return sorted_list

# 엑셀 파일 생성 함수
def create_result_wb(header, result):
    result_wb = Workbook()
    result_wb.active.title = 'result'
    result_ws = result_wb['result']

    # 엑셀 파일 상단 헤더 생성
    for cell in header:
        result_ws.cell(row=1, column=cell.col_idx, value=format_text(cell.value))

    # 엑셀 파일 내용 생성
    for i in range(len(result)):
        print(i)
        for cell in result[i]['row']:
            if(cell.col_idx == 1):
                result_ws.cell(row=i+2, column=cell.col_idx, value=i+1)
            else:
                result_ws.cell(row=i+2, column=cell.col_idx, value=format_text(cell.value))

    # 엑셀 파일 저장
    result_wb.save('result.xlsx')
    result_wb.close()

find_result1 = find_row_and_company_name('주식양수도.xlsx', '공시제목', '주식양수도')
find_result2 = find_row_and_company_name('전환사채.xlsx', '공시제목', '전환사채')
find_result3 = find_row_and_company_name('제3자배정.xlsx', '공시제목', '유상증자')

result = compare_company_names(find_result1['list'], find_result2['list'] + find_result3['list'])

create_result_wb(find_result1['header'], result)